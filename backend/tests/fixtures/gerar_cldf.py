"""Gera as fixtures XLSX da CLDF (rodar uma vez: uv run python tests/fixtures/gerar_cldf.py)."""
from datetime import datetime
from pathlib import Path

import openpyxl

PASTA = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.append([
    "NOME_PARLAMENTAR", "CPF_PARLAMENTAR", "NOME_PRESTADOR", "CNPJ_PRESTADOR",
    "CPF_PRESTADOR", "NR_COMPROVANTE", "DATA_COMPROVANTE", "VALOR_DESPESA",
    "CLASSIFICACAO", "OBSERVACOES",
])
ws.append(["Deputado Chico Vigilante", "111", "PAPELARIA ALFA LTDA",
           "01.111.111/0001-11", None, "10", datetime(2016, 3, 5), 250.0,
           "MATERIAL DE ESCRITÓRIO", None])
ws.append(["Deputado Chico Vigilante", "111", "POSTO BETA", None,
           "222.333.444-55", "11", datetime(2016, 3, 20), 100.0, None,
           "abastecimento"])
ws.append(["Deputada Jane Klebia", "222", "IMOBILIÁRIA GAMA",
           "02.222.222/0002-22", None, "12", datetime(2016, 4, 1), 3000.0,
           "Locação de imóvel", None])
ws.append(["Deputado Chico Vigilante", "111", "FORNECEDOR SEM DATA", None, None,
           "13", None, 999.0, None, None])
# 3 novas linhas com formatos reais
ws.append(["Deputada Jane Klebia", "222", "HOTEL XYZ",
           "03.333.333/0003-33", None, "14", datetime(2016, 5, 10), "R$ 5.000,00",
           "Hospedagem", None])
ws.append(["Deputado Chico Vigilante", "111", "TRANSPORTADORA ABC",
           "04.444.444/0004-44", None, "15", datetime(2016, 6, 15), "9,900,00",
           "Locação de veículos", None])
ws.append(["Deputada Jane Klebia", "222", "RESTAURANTE DEF", None,
           "555.666.777-88", "16", "15/05/2016", 700,
           "Alimentação", None])
wb.save(PASTA / "cldf_transacional.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Filtros aplicados:\nAno é 2025"] + [None] * 6)
ws.append([None] * 7)
ws.append(["ano", "mês", "deputado", "Imóvel", "Veículos", "Glosa", "totalVerbaGeral"])
ws.append([2025, "abr", "CHICO VIGILANTE", 1000, 500, 0, 1800])
ws.append([2025, "mai", "JANE KLEBIA", 2000, None, 100, 1900])
wb.save(PASTA / "cldf_pivo.xlsx")

# 2013: ordem diferente (DATA_COMPROVANTE antes de NR_COMPROVANTE)
wb = openpyxl.Workbook()
ws = wb.active
ws.append([
    "NOME_PARLAMENTAR", "CPF_PARLAMENTAR", "NOME_PRESTADOR", "CNPJ_PRESTADOR",
    "CPF_PRESTADOR", "DATA_COMPROVANTE", "NR_COMPROVANTE", "VALOR_DESPESA",
    "CLASSIFICACAO", "OBSERVACOES",
])
ws.append(["Deputado Chico Vigilante", "111", "PAPELARIA ALFA LTDA",
           "01.111.111/0001-11", None, datetime(2013, 3, 5), "10", 250.0,
           "MATERIAL DE ESCRITÓRIO", None])
wb.save(PASTA / "cldf_transacional_2013.xlsx")
print("fixtures geradas")
